"""
Statistics and Excel report generation service
"""
import io
from datetime import datetime, timedelta
from typing import Dict, Any, Optional

# Сначала импортируем базовые модули
from app.database import get_db_connection
from app.logging_config import get_logger

logger = get_logger(__name__)


def make_naive_datetime(dt_value) -> Optional[datetime]:
    """
    Преобразует datetime объект в naive (без timezone) для совместимости с Excel.
    
    Args:
        dt_value: datetime объект (может быть timezone-aware или naive), строка или None
    
    Returns:
        datetime: Naive datetime объект или None
    """
    if dt_value is None:
        return None
    
    # Если это строка, пытаемся распарсить
    if isinstance(dt_value, str):
        if not dt_value.strip():
            return None
        try:
            # Пробуем распарсить ISO формат
            if 'T' in dt_value or '+' in dt_value or 'Z' in dt_value:
                dt = datetime.fromisoformat(dt_value.replace('Z', '+00:00'))
            else:
                # Пробуем стандартные форматы
                for fmt in ['%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d']:
                    try:
                        dt = datetime.strptime(dt_value, fmt)
                        break
                    except ValueError:
                        continue
                else:
                    # Если не удалось распарсить, возвращаем None
                    logger.debug(f"⚠️ Could not parse datetime string: {dt_value}")
                    return None
        except (ValueError, AttributeError) as e:
            logger.debug(f"⚠️ Error parsing datetime: {e}")
            return None
    elif isinstance(dt_value, datetime):
        dt = dt_value
    else:
        # Неизвестный тип
        logger.debug(f"⚠️ Unknown datetime type: {type(dt_value)}")
        return None
    
    # Если datetime имеет timezone, преобразуем в naive
    if dt.tzinfo is not None:
        try:
            # Преобразуем в локальное время, затем убираем timezone
            dt = dt.astimezone(tz=None).replace(tzinfo=None)
        except (ValueError, OSError) as e:
            logger.warning(f"⚠️ Error converting timezone-aware datetime: {e}, using replace")
            # Если не удалось преобразовать, просто убираем timezone
            dt = dt.replace(tzinfo=None)
    
    return dt

# Опциональные импорты для графиков (после инициализации logger)
MATPLOTLIB_AVAILABLE = False
OPENPYXL_AVAILABLE = False

try:
    import matplotlib
    matplotlib.use('Agg')  # Non-GUI backend
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
    logger.debug("✅ matplotlib loaded successfully")
except ImportError as e:
    MATPLOTLIB_AVAILABLE = False
    logger.warning(f"⚠️ matplotlib not available, charts will be disabled: {e}")

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
    logger.debug("✅ openpyxl loaded successfully")
except ImportError as e:
    OPENPYXL_AVAILABLE = False
    logger.warning(f"⚠️ openpyxl not available, Excel export will be disabled: {e}")


def get_dashboard_statistics(user_role: str = 'admin') -> Dict[str, Any]:
    """
    Получить статистику для дашборда
    
    Args:
        user_role: Роль пользователя ('admin' или 'employee')
    
    Returns:
        Dict с метриками статистики
    """
    logger.info(f"📊 Generating dashboard statistics for role: {user_role}")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        stats = {}
        
        # Общее количество задач
        cur.execute("SELECT COUNT(*) as count FROM tasks")
        result = cur.fetchone()
        stats['total_tasks'] = result['count'] if result and 'count' in result else 0
        
        # Задачи по статусам
        cur.execute("""
            SELECT status, COUNT(*) as count
            FROM tasks 
            GROUP BY status
        """)
        status_results = cur.fetchall()
        # Правильно создаём словарь из результатов dict_factory
        status_counts = {row['status']: row['count'] for row in status_results}
        stats['by_status'] = {
            'pending': status_counts.get('pending', 0),
            'in_progress': status_counts.get('in_progress', 0),
            'partially_completed': status_counts.get('partially_completed', 0),
            'completed': status_counts.get('completed', 0),
            'rejected': status_counts.get('rejected', 0)
        }
        
        # Активные задачи (не завершённые)
        # Частично завершённые задачи считаются завершёнными
        stats['active_tasks'] = (
            stats['by_status']['pending'] + 
            stats['by_status']['in_progress']
        )
        
        # Задачи по приоритетам
        cur.execute("""
            SELECT priority, COUNT(*) as count
            FROM tasks 
            WHERE status NOT IN ('completed', 'partially_completed', 'rejected')
            GROUP BY priority
        """)
        priority_results = cur.fetchall()
        # Правильно создаём словарь из результатов dict_factory
        stats['by_priority'] = {row['priority']: row['count'] for row in priority_results} if priority_results else {}
        
        # Просроченные задачи
        cur.execute("""
            SELECT COUNT(*) as count 
            FROM tasks 
            WHERE due_date < datetime('now') 
            AND status NOT IN ('completed', 'partially_completed', 'rejected')
        """)
        result = cur.fetchone()
        stats['overdue_tasks'] = result['count'] if result and 'count' in result else 0
        
        # Задачи за сегодня (созданные)
        cur.execute("""
            SELECT COUNT(*) as count 
            FROM tasks 
            WHERE DATE(created_at) = DATE('now')
        """)
        result = cur.fetchone()
        stats['today_created'] = result['count'] if result and 'count' in result else 0
        
        # Завершённые за последние 7 дней
        cur.execute("""
            SELECT COUNT(*) as count 
            FROM tasks 
            WHERE status = 'completed'
            AND updated_at >= datetime('now', '-7 days')
        """)
        result = cur.fetchone()
        stats['completed_last_week'] = result['count'] if result and 'count' in result else 0
        
        # Топ исполнителей
        cur.execute("""
            SELECT u.username, u.first_name, u.last_name, COUNT(t.id) as task_count
            FROM tasks t
            JOIN users u ON t.assigned_to_id = u.id
            WHERE t.status = 'completed'
            GROUP BY u.username, u.first_name, u.last_name
            ORDER BY task_count DESC
            LIMIT 5
        """)
        stats['top_performers'] = cur.fetchall() or []
        
        logger.info(f"✅ Dashboard statistics generated: {stats['total_tasks']} total tasks")
        logger.debug(f"📊 Stats details: active={stats['active_tasks']}, overdue={stats['overdue_tasks']}, performers={len(stats['top_performers'])}")
        return stats
        
    except Exception as e:
        logger.error(f"❌ Error generating dashboard statistics: {e}", exc_info=True)
        return {}
    finally:
        cur.close()
        conn.close()


def generate_excel_report(report_type: str = 'full') -> io.BytesIO:
    """
    Генерация Excel отчёта с графиками
    
    Args:
        report_type: Тип отчёта ('full', 'status', 'priority', 'users')
    
    Returns:
        BytesIO объект с Excel файлом
    
    Raises:
        ImportError: Если openpyxl не установлен
        Exception: При ошибках генерации отчёта
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl не установлен. Установите его командой: pip install openpyxl\n"
            "Для графиков также нужен matplotlib: pip install matplotlib"
        )
    
    logger.info(f"📊 Generating Excel report: {report_type}")
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика задач"
        
        # Заголовок
        ws['A1'] = "Отчёт по задачам"
        ws['A1'].font = Font(size=16, bold=True)
        # Используем naive datetime для строки
        now_naive = datetime.now()
        if now_naive.tzinfo is not None:
            now_naive = now_naive.replace(tzinfo=None)
        ws['A2'] = f"Дата создания: {now_naive.strftime('%d.%m.%Y %H:%M')}"
        ws['A2'].font = Font(size=10, italic=True)
        
        stats = get_dashboard_statistics()
        
        if not stats:
            logger.warning("⚠️ No statistics data available")
            ws['A4'] = "Нет данных для отчёта"
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output
        
        # Общая статистика
        ws['A4'] = "Общая статистика"
        ws['A4'].font = Font(size=14, bold=True)
        ws['A4'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A4'].font = Font(size=14, bold=True, color="FFFFFF")
        
        row = 5
        ws[f'A{row}'] = "Всего задач:"
        ws[f'B{row}'] = stats.get('total_tasks', 0)
        ws[f'B{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Активных задач:"
        ws[f'B{row}'] = stats.get('active_tasks', 0)
        
        row += 1
        ws[f'A{row}'] = "Завершено:"
        ws[f'B{row}'] = stats.get('by_status', {}).get('completed', 0)
        
        row += 1
        ws[f'A{row}'] = "Просрочено:"
        ws[f'B{row}'] = stats.get('overdue_tasks', 0)
        ws[f'B{row}'].font = Font(color="FF0000", bold=True)
        
        # Статистика по статусам
        row += 2
        ws[f'A{row}'] = "Задачи по статусам"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
        
        row += 1
        status_labels = {
            'pending': 'Ожидает',
            'in_progress': 'В работе',
            'partially_completed': 'Частично завершена',
            'completed': 'Завершена',
            'rejected': 'Отклонена'
        }
        
        by_status = stats.get('by_status', {})
        for status_key, label in status_labels.items():
            ws[f'A{row}'] = label
            ws[f'B{row}'] = by_status.get(status_key, 0)
            row += 1
        
        # Топ исполнителей
        if stats.get('top_performers'):
            row += 1
            ws[f'A{row}'] = "Топ исполнителей"
            ws[f'A{row}'].font = Font(size=12, bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            ws[f'A{row}'].font = Font(size=12, bold=True, color="000000")
            
            row += 1
            ws[f'A{row}'] = "Исполнитель"
            ws[f'B{row}'] = "Завершено задач"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            
            row += 1
            for performer in stats['top_performers']:
                username = performer.get('username', 'Неизвестно')
                first_name = performer.get('first_name')
                last_name = performer.get('last_name')
                count = performer.get('task_count', 0)
                
                # Форматируем имя исполнителя
                if first_name or last_name:
                    user_display = f"{first_name or ''} {last_name or ''}".strip() + f" (@{username})"
                else:
                    user_display = f"@{username}"
                
                ws[f'A{row}'] = user_display
                ws[f'B{row}'] = count
                row += 1
        
        # Автоматическая ширина столбцов
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
        # Создание графиков с matplotlib (если доступен)
        if MATPLOTLIB_AVAILABLE and OPENPYXL_AVAILABLE:
            try:
                # График по статусам
                fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
                
                # Круговая диаграмма по статусам
                status_data = [v for v in by_status.values() if v > 0]
                status_labels_filtered = [status_labels[k] for k, v in by_status.items() if v > 0]
                colors = ['#FFA500', '#4169E1', '#FFD700', '#32CD32', '#FF6347']
                
                if status_data:
                    ax1.pie(status_data, labels=status_labels_filtered, autopct='%1.1f%%', colors=colors)
                    ax1.set_title('Распределение задач по статусам', fontsize=14, fontweight='bold')
                else:
                    ax1.text(0.5, 0.5, 'Нет данных', ha='center', va='center', fontsize=14)
                
                # Столбчатая диаграмма приоритетов
                by_priority = stats.get('by_priority', {})
                priority_labels = list(by_priority.keys())
                priority_values = list(by_priority.values())
                
                if priority_labels and priority_values:
                    ax2.bar(priority_labels, priority_values, color=['#FF4444', '#FF8800', '#FFDD44', '#88DD44'])
                    ax2.set_title('Задачи по приоритетам (активные)', fontsize=14, fontweight='bold')
                    ax2.set_xlabel('Приоритет')
                    ax2.set_ylabel('Количество задач')
                else:
                    ax2.text(0.5, 0.5, 'Нет данных', ha='center', va='center', fontsize=14)
                
                plt.tight_layout()
                
                # Сохранение графика в BytesIO
                img_stream = io.BytesIO()
                plt.savefig(img_stream, format='png', dpi=100, bbox_inches='tight')
                img_stream.seek(0)
                plt.close(fig)
                
                # Вставка изображения в Excel (на новый лист)
                try:
                    ws_charts = wb.create_sheet(title="Графики")
                    img = XLImage(img_stream)
                    ws_charts.add_image(img, 'A1')
                    logger.info("✅ Excel charts generated successfully")
                except Exception as img_error:
                    logger.warning(f"⚠️ Could not add image to Excel: {img_error}")
                    # Продолжаем без изображения
                
            except Exception as e:
                logger.error(f"❌ Error generating charts: {e}", exc_info=True)
        else:
            logger.debug("⚠️ Charts disabled: matplotlib or openpyxl not available")
        
        # Добавляем детальные таблицы по исполнителям
        conn = get_db_connection()
        cur = conn.cursor()
        
        try:
            # Лист с выполненными задачами по исполнителям
            ws_completed = wb.create_sheet(title="Выполненные задачи")
            ws_completed['A1'] = "Выполненные задачи по исполнителям"
            ws_completed['A1'].font = Font(size=14, bold=True)
            ws_completed['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            ws_completed['A1'].font = Font(size=14, bold=True, color="FFFFFF")
            
            # Заголовки таблицы
            ws_completed['A3'] = "Исполнитель"
            ws_completed['B3'] = "ID Задачи"
            ws_completed['C3'] = "Название"
            ws_completed['D3'] = "Приоритет"
            ws_completed['E3'] = "Дата завершения"
            
            for col in ['A3', 'B3', 'C3', 'D3', 'E3']:
                ws_completed[col].font = Font(bold=True)
                ws_completed[col].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Получаем выполненные задачи
            cur.execute("""
                SELECT 
                    u.username,
                    u.first_name,
                    u.last_name,
                    t.id,
                    t.title,
                    t.priority,
                    t.updated_at
                FROM tasks t
                JOIN users u ON t.assigned_to_id = u.id
                WHERE t.status = 'completed'
                ORDER BY u.username, t.updated_at DESC
            """)
            
            completed_tasks = cur.fetchall()
            row_completed = 4
            for task in completed_tasks:
                username = task.get('username', 'Неизвестно')
                first_name = task.get('first_name')
                last_name = task.get('last_name')
                task_id = task.get('id', 0)
                title = task.get('title', '')[:50]  # Ограничение длины
                priority = task.get('priority', '')
                updated_at = task.get('updated_at', '')
                
                # Форматируем имя исполнителя
                if first_name or last_name:
                    user_display = f"{first_name or ''} {last_name or ''}".strip() + f" (@{username})"
                else:
                    user_display = f"@{username}"
                
                ws_completed[f'A{row_completed}'] = user_display
                ws_completed[f'B{row_completed}'] = task_id
                ws_completed[f'C{row_completed}'] = title
                ws_completed[f'D{row_completed}'] = priority
                # Преобразуем datetime в naive для Excel
                naive_dt = make_naive_datetime(updated_at)
                ws_completed[f'E{row_completed}'] = naive_dt if naive_dt else ''
                row_completed += 1
            
            # Автоширина столбцов
            ws_completed.column_dimensions['A'].width = 20
            ws_completed.column_dimensions['B'].width = 10
            ws_completed.column_dimensions['C'].width = 40
            ws_completed.column_dimensions['D'].width = 12
            ws_completed.column_dimensions['E'].width = 18
            
            logger.info(f"✅ Added {len(completed_tasks)} completed tasks to report")
            
            # Лист с просроченными задачами
            ws_overdue = wb.create_sheet(title="Просроченные задачи")
            ws_overdue['A1'] = "Просроченные задачи по исполнителям"
            ws_overdue['A1'].font = Font(size=14, bold=True)
            ws_overdue['A1'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            ws_overdue['A1'].font = Font(size=14, bold=True, color="FFFFFF")
            
            # Заголовки таблицы
            ws_overdue['A3'] = "Исполнитель"
            ws_overdue['B3'] = "ID Задачи"
            ws_overdue['C3'] = "Название"
            ws_overdue['D3'] = "Приоритет"
            ws_overdue['E3'] = "Срок выполнения"
            ws_overdue['F3'] = "Статус"
            ws_overdue['G3'] = "Просрочено дней"
            
            for col in ['A3', 'B3', 'C3', 'D3', 'E3', 'F3', 'G3']:
                ws_overdue[col].font = Font(bold=True)
                ws_overdue[col].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            
            # Получаем просроченные задачи
            cur.execute("""
                SELECT 
                    u.username,
                    u.first_name,
                    u.last_name,
                    t.id,
                    t.title,
                    t.priority,
                    t.due_date,
                    t.status,
                    CAST((julianday('now') - julianday(t.due_date)) AS INTEGER) as days_overdue
                FROM tasks t
                JOIN users u ON t.assigned_to_id = u.id
                WHERE t.due_date < datetime('now') 
                AND t.status NOT IN ('completed', 'partially_completed', 'rejected')
                ORDER BY t.due_date ASC
            """)
            
            overdue_tasks = cur.fetchall()
            row_overdue = 4
            for task in overdue_tasks:
                username = task.get('username', 'Неизвестно')
                first_name = task.get('first_name')
                last_name = task.get('last_name')
                task_id = task.get('id', 0)
                title = task.get('title', '')[:50]
                priority = task.get('priority', '')
                due_date = task.get('due_date', '')
                status = task.get('status', '')
                days_overdue = task.get('days_overdue', 0)
                
                # Форматируем имя исполнителя
                if first_name or last_name:
                    user_display = f"{first_name or ''} {last_name or ''}".strip() + f" (@{username})"
                else:
                    user_display = f"@{username}"
                
                ws_overdue[f'A{row_overdue}'] = user_display
                ws_overdue[f'B{row_overdue}'] = task_id
                ws_overdue[f'C{row_overdue}'] = title
                ws_overdue[f'D{row_overdue}'] = priority
                # Преобразуем datetime в naive для Excel
                naive_dt = make_naive_datetime(due_date)
                ws_overdue[f'E{row_overdue}'] = naive_dt if naive_dt else ''
                ws_overdue[f'F{row_overdue}'] = status
                ws_overdue[f'G{row_overdue}'] = int(days_overdue) if days_overdue else 0
                
                # Красный цвет для строк с высокой просрочкой
                if days_overdue and days_overdue > 7:
                    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                        ws_overdue[f'{col}{row_overdue}'].fill = PatternFill(
                            start_color="FFB3B3", end_color="FFB3B3", fill_type="solid"
                        )
                
                row_overdue += 1
            
            # Автоширина столбцов
            ws_overdue.column_dimensions['A'].width = 20
            ws_overdue.column_dimensions['B'].width = 10
            ws_overdue.column_dimensions['C'].width = 40
            ws_overdue.column_dimensions['D'].width = 12
            ws_overdue.column_dimensions['E'].width = 15
            ws_overdue.column_dimensions['F'].width = 15
            ws_overdue.column_dimensions['G'].width = 15
            
            logger.info(f"✅ Added {len(overdue_tasks)} overdue tasks to report")
            
        except Exception as e:
            logger.error(f"❌ Error adding detailed task tables: {e}", exc_info=True)
            # Продолжаем даже если детальные таблицы не добавились
        finally:
            cur.close()
            conn.close()
        
        # Сохранение в BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        file_size = output.getbuffer().nbytes
        logger.info(f"✅ Excel report generated successfully: {file_size} bytes")
        
        if file_size == 0:
            raise Exception("Сгенерированный файл пуст")
        
        return output
        
    except ImportError:
        # Пробрасываем ImportError как есть
        raise
    except Exception as e:
        logger.error(f"❌ Error in generate_excel_report: {e}", exc_info=True)
        raise Exception(f"Ошибка при генерации Excel отчёта: {str(e)}")
